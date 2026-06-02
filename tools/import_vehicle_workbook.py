import datetime as dt
import math
import re
import uuid
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
INPUT = Path(r"C:\Users\nioon\Downloads\車輛總表 的副本.xlsx")
OUTPUT = ROOT / "supabase" / "migrations" / "20260602010000_import_vehicle_workbook.sql"
NS = uuid.UUID("58d84afa-6f43-4b9d-81d4-5441f31d8870")


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    if text in {"nan", "NaN", "#REF!", "#VALUE!", "#DIV/0!"}:
        return ""
    if text.endswith(".0") and re.fullmatch(r"-?\d+\.0", text):
        text = text[:-2]
    return text.strip()


def num(value):
    text = clean(value).replace(",", "")
    if not text:
        return "0"
    try:
        return str(float(text))
    except ValueError:
        return "0"


def intnum(value):
    text = clean(value).replace(",", "")
    if not text:
        return "0"
    try:
        return str(int(float(text)))
    except ValueError:
        return "0"


def date_or_null(value):
    text = clean(value)
    if not text or text.startswith("1900-01"):
        return "null"
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        year = int(text[:4])
        if year < 2000 or year > 2100:
            return "null"
        return sql(text)
    return "null"


def sql(value):
    if value is None:
        return "null"
    return "'" + str(value).replace("'", "''") + "'"


def qnum(value):
    return num(value)


def vid(plate):
    return str(uuid.uuid5(NS, f"vehicle:{plate.upper()}"))


def rid(prefix, *parts):
    key = ":".join(clean(p) for p in parts if clean(p))
    return str(uuid.uuid5(NS, f"{prefix}:{key}"))


def get_sheet_rows(wb, name, header_row):
    ws = wb[name]
    seen = {}
    headers = []
    for c in ws[header_row]:
        header = clean(c.value)
        if header:
            seen[header] = seen.get(header, 0) + 1
            if seen[header] > 1:
                header = f"{header}.{seen[header] - 1}"
        headers.append(header)
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        data = {headers[i] or f"col_{i}": row[i] if i < len(row) else None for i in range(len(headers))}
        rows.append(data)
    return headers, rows


def unique_vendor(vendors, category, name):
    name = clean(name)
    if name and name not in {"V", "X", "-", "無"}:
        vendors.add((category, name))


def main():
    wb = openpyxl.load_workbook(INPUT, read_only=True, data_only=True)
    vehicles = {}
    vendors = set()

    _, rows = get_sheet_rows(wb, "車輛總表", 1)
    for r in rows:
        plate = clean(r.get("車號"))
        if not plate:
            continue
        vehicles[plate] = {
            "id": vid(plate),
            "plate_no": plate,
            "model": clean(r.get("車輛款式")),
            "brand": clean(r.get("車輛品牌")),
            "body_color": clean(r.get("車身顏色")),
            "fuel_type": clean(r.get("油品")),
            "manufacture_date": date_or_null(r.get("出廠日期")),
            "deposit_date": date_or_null(r.get("付訂日期")),
            "final_payment_date": date_or_null(r.get("尾款日期")),
            "license_plate_date": date_or_null(r.get("領牌日期")),
            "delivery_date": date_or_null(r.get("交車日期")),
            "purchase_total_cost": qnum(r.get("購置總成本")),
            "dealer": clean(r.get("車輛經銷商")),
            "loan_bank": clean(r.get("貸款銀行")),
            "insurance_company": clean(r.get("保險公司")),
            "original_plate_owner": clean(r.get("原鐵牌所屬")),
            "status": clean(r.get("目前狀態")) or "其他",
            "withholding_dealer": clean(r.get("代扣車商")),
            "withholding_person": clean(r.get("代扣人")),
            "notes": clean(r.get("備註")),
            "notes2": clean(r.get("備註2")),
            "purchase_subsidy_enabled": "true" if clean(r.get("購車補貼")) else "false",
            "purchase_subsidy_amount": qnum(r.get("購車補貼.1")),
            "fleet_name": "亞菲得車隊",
        }
        unique_vendor(vendors, "車商", r.get("車輛經銷商"))
        unique_vendor(vendors, "貸款銀行", r.get("貸款銀行"))
        unique_vendor(vendors, "保險公司", r.get("保險公司"))
        unique_vendor(vendors, "鐵牌所屬", r.get("原鐵牌所屬"))
        unique_vendor(vendors, "代扣車商", r.get("代扣車商"))

    # Supplement active use, inspection, and insurance expiry from the operational sheet.
    _, self_rows = get_sheet_rows(wb, "自用車輛管理", 2)
    for r in self_rows:
        plate = clean(r.get("車牌號碼"))
        if not plate:
            continue
        item = vehicles.setdefault(plate, {"id": vid(plate), "plate_no": plate, "fleet_name": "亞菲得車隊"})
        item.setdefault("model", clean(r.get("車輛款式")))
        item.setdefault("body_color", clean(r.get("車身顏色")))
        item.setdefault("fuel_type", clean(r.get("油品")))
        item["current_usage"] = clean(r.get("目前用途")) or item.get("current_usage", "")
        item["original_plate_owner"] = clean(r.get("原鐵牌所屬")) or item.get("original_plate_owner", "")
        item["insurance_company"] = clean(r.get("保險公司")) or item.get("insurance_company", "")
        item["insurance_expiry"] = date_or_null(r.get("任意險到期日"))
        item["next_inspection_date"] = date_or_null(r.get("下次驗車日期"))
        unique_vendor(vendors, "目前用途", r.get("目前用途"))
        unique_vendor(vendors, "鐵牌所屬", r.get("原鐵牌所屬"))
        unique_vendor(vendors, "保險公司", r.get("保險公司"))

    insurance = []
    _, ins_rows = get_sheet_rows(wb, "車輛保險管理", 2)
    for r in ins_rows:
        plate = clean(r.get("車牌號碼"))
        if not plate:
            continue
        vehicles.setdefault(plate, {"id": vid(plate), "plate_no": plate, "fleet_name": "亞菲得車隊"})
        insurance.append({
            "id": rid("insurance", plate, r.get("投保日期"), r.get("到期日期"), r.get("投保公司")),
            "vehicle_id": vid(plate),
            "plate_no": plate,
            "compulsory_start_date": date_or_null(r.get("投保日期")),
            "compulsory_end_date": date_or_null(r.get("到期日期")),
            "compulsory_company": clean(r.get("投保公司")),
            "optional_start_date": date_or_null(r.get("投保日期.1")),
            "optional_end_date": date_or_null(r.get("到期日期.1")),
            "optional_company": clean(r.get("投保公司.1")),
            "broker": clean(r.get("業務人員.1")) or clean(r.get("業務人員")),
            "total_premium": str(float(num(r.get("投保費用/年"))) + float(num(r.get("投保費用/年.1")))),
            "notes": clean(r.get("備註")),
        })
        unique_vendor(vendors, "保險公司", r.get("投保公司"))
        unique_vendor(vendors, "保險公司", r.get("投保公司.1"))
        unique_vendor(vendors, "貸款銀行", r.get("貸款銀行"))

    maintenance = []
    ws = wb["車輛保養排程"]
    headers = [c.value for c in ws[2]]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    last_plate = None
    for idx, row in enumerate(rows):
        plate = clean(row[1] if len(row) > 1 else "")
        if plate:
            last_plate = plate
        if not plate:
            continue
        vehicles.setdefault(plate, {"id": vid(plate), "plate_no": plate, "fleet_name": "亞菲得車隊"})
        vendor = clean(row[5] if len(row) > 5 else "")
        unique_vendor(vendors, "保養廠", vendor)
        event_dates = []
        for col in range(6, len(headers)):
            h = headers[col]
            if not isinstance(h, (dt.date, dt.datetime)):
                continue
            value = clean(row[col] if col < len(row) else "")
            if not value:
                continue
            cost = ""
            if idx + 1 < len(rows):
                cost = clean(rows[idx + 1][col] if col < len(rows[idx + 1]) else "")
            event_dates.append((h.date() if isinstance(h, dt.datetime) else h, value, cost))
        for pos, (d, detail, cost) in enumerate(event_dates):
            next_date = event_dates[pos + 1][0].isoformat() if pos + 1 < len(event_dates) else ""
            maintenance.append({
                "id": rid("maintenance", plate, d.isoformat(), detail),
                "vehicle_id": vid(plate),
                "plate_no": plate,
                "service_date": d.isoformat(),
                "service_month": d.isoformat()[:7],
                "service_location": clean(row[3] if len(row) > 3 else ""),
                "vendor": vendor,
                "mileage": intnum(clean(re.search(r"(\d+)\s*萬", detail).group(1) + "0000") if re.search(r"(\d+)\s*萬", detail) else "0"),
                "cost": qnum(cost),
                "items": detail,
                "next_service_date": next_date,
            })

    tires = []
    ws = wb["車輛輪胎排程"]
    tire_headers = [clean(c.value) for c in ws[2]]
    for row in ws.iter_rows(min_row=3, values_only=True):
        plate = clean(row[1] if len(row) > 1 else "")
        if not plate:
            continue
        vehicles.setdefault(plate, {"id": vid(plate), "plate_no": plate, "fleet_name": "亞菲得車隊"})
        base_vendor = clean(row[4] if len(row) > 4 else "")
        unique_vendor(vendors, "保養廠", base_vendor)
        for col, header in enumerate(tire_headers):
            if not header.startswith("更換日期"):
                continue
            d = row[col] if col < len(row) else None
            if not isinstance(d, (dt.date, dt.datetime)):
                continue
            vendor = clean(row[col + 1] if col + 1 < len(row) else "") or base_vendor
            detail = clean(row[col + 2] if col + 2 < len(row) else "")
            mileage = clean(row[col + 3] if col + 3 < len(row) else "")
            unique_vendor(vendors, "保養廠", vendor)
            tires.append({
                "id": rid("tire", plate, d.isoformat(), vendor, detail, mileage),
                "vehicle_id": vid(plate),
                "plate_no": plate,
                "replacement_date": (d.date() if isinstance(d, dt.datetime) else d).isoformat(),
                "mileage": intnum(mileage),
                "tire_type": detail,
                "vendor": vendor,
                "details": detail,
            })

    lines = [
        "-- Generated from 車輛總表 的副本.xlsx.",
        "begin;",
    ]
    for category, name in sorted(vendors):
        lines.append(
            "insert into public.vendor_options (id, category, name) values "
            f"({sql(rid('vendor', category, name))}, {sql(category)}, {sql(name)}) "
            "on conflict (id) do update set category = excluded.category, name = excluded.name, updated_at = now();"
        )

    vehicle_cols = [
        "id", "plate_no", "model", "brand", "body_color", "fuel_type", "manufacture_date",
        "deposit_date", "final_payment_date", "license_plate_date", "delivery_date", "purchase_total_cost",
        "dealer", "loan_bank", "insurance_company", "original_plate_owner", "status", "current_usage",
        "withholding_dealer", "withholding_person", "notes", "notes2", "purchase_subsidy_enabled",
        "purchase_subsidy_amount", "fleet_name", "insurance_expiry", "next_inspection_date"
    ]
    for v in vehicles.values():
        vals = []
        for col in vehicle_cols:
            if col in {"manufacture_date", "deposit_date", "final_payment_date", "license_plate_date", "delivery_date", "insurance_expiry", "next_inspection_date"}:
                vals.append(v.get(col, "null"))
            elif col in {"purchase_total_cost", "purchase_subsidy_amount"}:
                vals.append(v.get(col, "0"))
            elif col == "purchase_subsidy_enabled":
                vals.append(v.get(col, "false"))
            else:
                vals.append(sql(v.get(col, "")))
        updates = ", ".join(f"{c}=excluded.{c}" for c in vehicle_cols if c not in {"id", "plate_no"})
        lines.append(f"insert into public.vehicles ({', '.join(vehicle_cols)}) values ({', '.join(vals)}) on conflict (plate_no) do update set {updates}, updated_at = now();")

    for item in insurance:
        lines.append(
            "insert into public.insurance_records "
            "(id, vehicle_id, compulsory_company, optional_company, compulsory_start_date, compulsory_end_date, optional_start_date, optional_end_date, broker, total_premium, notes) values "
            f"({sql(item['id'])}, (select id from public.vehicles where plate_no = {sql(item['plate_no'])} limit 1), {sql(item['compulsory_company'])}, {sql(item['optional_company'])}, {item['compulsory_start_date']}, {item['compulsory_end_date']}, {item['optional_start_date']}, {item['optional_end_date']}, {sql(item['broker'])}, {item['total_premium']}, {sql(item['notes'])}) "
            "on conflict (id) do update set compulsory_company=excluded.compulsory_company, optional_company=excluded.optional_company, compulsory_start_date=excluded.compulsory_start_date, compulsory_end_date=excluded.compulsory_end_date, optional_start_date=excluded.optional_start_date, optional_end_date=excluded.optional_end_date, broker=excluded.broker, total_premium=excluded.total_premium, notes=excluded.notes, updated_at=now();"
        )

    for item in maintenance:
        lines.append(
            "insert into public.maintenance_records "
            "(id, vehicle_id, service_date, service_month, service_location, vendor, mileage, cost, items, next_service_date) values "
            f"({sql(item['id'])}, (select id from public.vehicles where plate_no = {sql(item['plate_no'])} limit 1), {sql(item['service_date'])}, {sql(item['service_month'])}, {sql(item['service_location'])}, {sql(item['vendor'])}, {item['mileage']}, {item['cost']}, {sql(item['items'])}, {sql(item['next_service_date']) if item['next_service_date'] else 'null'}) "
            "on conflict (id) do update set service_date=excluded.service_date, service_month=excluded.service_month, service_location=excluded.service_location, vendor=excluded.vendor, mileage=excluded.mileage, cost=excluded.cost, items=excluded.items, next_service_date=excluded.next_service_date, updated_at=now();"
        )

    for item in tires:
        lines.append(
            "insert into public.tire_records "
            "(id, vehicle_id, replacement_date, mileage, tire_type, vendor, details) values "
            f"({sql(item['id'])}, (select id from public.vehicles where plate_no = {sql(item['plate_no'])} limit 1), {sql(item['replacement_date'])}, {item['mileage']}, {sql(item['tire_type'])}, {sql(item['vendor'])}, {sql(item['details'])}) "
            "on conflict (id) do update set replacement_date=excluded.replacement_date, mileage=excluded.mileage, tire_type=excluded.tire_type, vendor=excluded.vendor, details=excluded.details, updated_at=now();"
        )

    lines.append("commit;")
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"vehicles={len(vehicles)} vendors={len(vendors)} insurance={len(insurance)} maintenance={len(maintenance)} tires={len(tires)}")
    print(OUTPUT)


if __name__ == "__main__":
    main()
